#!/usr/bin/env python3
"""
update_dashboard.py  ·  1010 Church · Container Dashboard Updater
──────────────────────────────────────────────────────────────────
Double-click to run  (or:  python update_dashboard.py)

What this does:
  1. Reads the Excel file and extracts fresh data
  2. Opens the existing HTML file (keeping ALL layout/design intact)
  3. Replaces ONLY the data block inside the HTML
  4. Runs git add → commit → push so GitHub Pages auto-updates

The HTML file is the permanent source of truth for design.
This script only touches the data — nothing else.
"""

import os, sys, re, subprocess
from datetime import datetime

# ── AUTO-INSTALL openpyxl IF MISSING ─────────────────────────
try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
except ImportError:
    print("openpyxl not found — installing now...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.utils.datetime import from_excel

# ── CONFIG ────────────────────────────────────────────────────
SITE_DIR    = r"C:\Users\Ewan\OneDrive - Studio Snaidero Chicago\Shared\1010_Logistics Site"
EXCEL_FILE  = os.path.join(SITE_DIR, "10C_Logistcs Output.xlsm")
OUTPUT_HTML = os.path.join(SITE_DIR, "10C_Container_Dashboard.html")
AUTO_GIT    = True   # set to False to skip the git push step

# ── SHEET NAME VARIANTS (uses first match) ────────────────────
SH_CTN_VARIANTS = ["10C_CTNR Schedule (spill)", "1010C_Container Table"]
SH_BLD_VARIANTS = ["10C Building Matrix (status)", "10C_BUILDING MATRIX"]
SH_ROJ_VARIANTS = ["1010 Church ROJ Dates"]   # optional

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def to_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)) and v > 40000:
        try:
            return from_excel(int(v))
        except Exception:
            return None
    return None

def fmt_date(v, year=False):
    d = to_dt(v)
    if not d:
        return "—"
    s = f"{d.strftime('%b')} {d.day}"
    return s + f", {d.year}" if year else s

def clean(v):
    if v is None:
        return "—"
    s = str(v).strip()
    return s if s else "—"

STATUS_NORM = {
    "PROJ":"PROJ","LDG":"LDG","LDD":"LDD","ENR":"ENR","CHS":"CHS","RAIL":"RAIL","INYRD":"INYRD","D":"D",
    "LOADING":"LDG","LOADED":"LDD","EN ROUTE":"ENR","ENROUTE":"ENR","SEA FREIGHT":"ENR",
    "PORT":"CHS","PORT (CHS)":"CHS","IN PORT":"CHS","CHS PORT":"CHS",
    "ON RAIL":"RAIL","RAIL FREIGHT":"RAIL","ON RAIL FREIGHT":"RAIL","ONRAIL":"RAIL","RFT":"RAIL","ONR":"RAIL",
    "IN YARD":"INYRD","IN YARD (NSH)":"INYRD","NHS":"INYRD","NSH":"INYRD","YARD":"INYRD","PULLED":"INYRD","IYD":"INYRD","PLD":"INYRD",
    "DELIVERED":"D","DELIVERY":"D",
}
UNKNOWN_STATUSES = set()

def norm_status(raw):
    """Map a raw Excel status to a known code. Unknown codes pass through
    verbatim (so they stay visible on the dashboard) and get reported."""
    s = str(raw).strip().upper()
    if not s or s in ("—", "--", "NONE"):
        return "PROJ"
    if s in STATUS_NORM:
        return STATUS_NORM[s]
    UNKNOWN_STATUSES.add(s)
    return s

def js_str(v):
    if isinstance(v, str):
        return "'" + v.replace("\\", "\\\\").replace("'", "\\'") + "'"
    if isinstance(v, bool):
        return "true" if v else "false"
    if v is None:
        return "'—'"
    return repr(v)

# ─────────────────────────────────────────────────────────────
# LOAD WORKBOOK
# ─────────────────────────────────────────────────────────────

print(f"\n{'─'*56}")
print(f"  1010 Church · Dashboard Updater")
print(f"{'─'*56}")

if not os.path.exists(EXCEL_FILE):
    sys.exit(
        f"\nERROR: Excel file not found:\n  {EXCEL_FILE}\n"
        f"Check the path in the CONFIG section at the top of this script.\n"
    )

if not os.path.exists(OUTPUT_HTML):
    sys.exit(
        f"\nERROR: Dashboard HTML not found:\n  {OUTPUT_HTML}\n"
        f"Make sure 10C_Container_Dashboard.html is in:\n  {SITE_DIR}\n"
    )

print(f"\nReading  {os.path.basename(EXCEL_FILE)} …")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

def find_sheet(variants, required=True):
    for name in variants:
        if name in wb.sheetnames:
            return name
    if required:
        sys.exit(f"\nERROR: None of these sheets found: {variants}\n"
                 f"Available sheets: {wb.sheetnames}")
    return None

SH_CTN = find_sheet(SH_CTN_VARIANTS, required=True)
SH_BLD = find_sheet(SH_BLD_VARIANTS, required=True)
SH_ROJ = find_sheet(SH_ROJ_VARIANTS, required=False)
print(f"  Using sheets: '{SH_CTN}', '{SH_BLD}', ROJ: '{SH_ROJ or 'not found'}'")

# ─────────────────────────────────────────────────────────────
# 1.  CONTAINER TABLE
# ─────────────────────────────────────────────────────────────

ws = wb[SH_CTN]
all_rows = list(ws.iter_rows(values_only=True))

hdr_row, hdr = None, None
for i, row in enumerate(all_rows):
    if any(v and "STATUS" in str(v).upper() and "CONT" in str(v).upper() for v in row):
        hdr_row = i
        hdr = [str(v).strip() if v else "" for v in row]
        break

if hdr is None:
    sys.exit(f"ERROR: Could not find header row in '{SH_CTN}'.")

def C(*names):
    for n in names:
        try:
            return hdr.index(n)
        except ValueError:
            pass
    nl = names[0].lower()
    for i, h in enumerate(hdr):
        if nl in h.lower():
            return i
    return None

i_status   = C("CONT. STATUS", "CONT.STATUS", "STATUS")
i_num      = C("CONT. #", "CONT.#", "CTN #")
i_week     = C("WEEK")
i_floor    = C("FLOORS", "FLOOR")
i_qty      = C("UNIT Q.ty", "UNIT Q.TY", "UNIT QTY")
i_ordered  = C("ORDERED")
i_vessel   = C("VESSEL")
i_kitchens = C("Kitchens", "KITCHENS", "Kitchen")
i_v1       = C("Vanity 1", "VANITY 1", "V1")
i_v2       = C("Vanity 2", "VANITY 2", "V2")
i_load     = C("LOAD DATE")
i_ship     = C("SHIP DATE")
i_chs      = C("PORT ARRIVAL DATE/CHS", "PORT ARRIVAL DATE", "ARRIVAL DATE/CHS", "ETA CHS", "CHS")
i_rail     = C("RAIL DATE")
i_nash     = C("ARRIVAL DATE (Nashville)", "NASHVILLE DATE", "ETA NASH", "ARRIVAL DATE")
i_del      = C("DELIVERY DATE", "DELIVERED DATE", "DELIVERY")
i_confirm  = C("CONFIRM")

print(f"  Container table header on row {hdr_row+1}")

CONTAINERS = []
for row in all_rows[hdr_row + 1:]:
    if not row or all(v is None for v in row):
        continue
    num_raw = row[i_num] if i_num is not None else None
    if not isinstance(num_raw, (int, float)):
        continue
    num = int(num_raw)
    if num < 1 or num > 30:
        continue

    week_raw = row[i_week] if i_week is not None else None
    if isinstance(week_raw, (int, float)):
        week = f"Wk {int(week_raw)}"
    else:
        ws_str = clean(week_raw)
        week = ("Wk " + ws_str.replace("Week ", "")) if "Week" in ws_str else ws_str

    vessel = clean(row[i_vessel] if i_vessel is not None else None)
    if vessel in ("—", "--", "None"):
        vessel = "—"

    # Normalize status from Excel to known status codes
    raw_st = clean(row[i_status]).strip().upper() if i_status is not None else "PROJ"
    status = norm_status(raw_st)

    CONTAINERS.append({
        "num":      num,
        "status":   status,
        "week":     week,
        "floor":    int(row[i_floor])     if (i_floor   is not None and isinstance(row[i_floor],   (int, float))) else 0,
        "unitQty":  int(row[i_qty])       if (i_qty     is not None and isinstance(row[i_qty],     (int, float))) else 0,
        "ordered":  bool(row[i_ordered])  if i_ordered  is not None else False,
        "vessel":   vessel,
        "kitchens": clean(row[i_kitchens]) if i_kitchens is not None else "—",
        "v1":       clean(row[i_v1])       if i_v1       is not None else "—",
        "v2":       clean(row[i_v2])       if i_v2       is not None else "—",
        "loadDate": fmt_date(row[i_load])  if i_load is not None else "—",
        "shipDate": fmt_date(row[i_ship])  if i_ship is not None else "—",
        "etaChs":   fmt_date(row[i_chs])   if i_chs  is not None else "—",
        "railDate": fmt_date(row[i_rail])  if i_rail is not None else "—",
        "etaNash":  fmt_date(row[i_nash])  if i_nash is not None else "—",
        "delivery": fmt_date(row[i_del], year=True) if i_del is not None else "—",
        "confirmed": bool(row[i_confirm]) if (i_confirm is not None and isinstance(row[i_confirm], bool)) else False,
        "code":     "—",
        "delivTime":"—",
    })

CONTAINERS.sort(key=lambda c: c["num"])

# Append special containers (BH, ECT, WCT) — always at the end
CONTAINERS += [
    {"num":"BH",  "status":"PROJ", "week":"Wk 26", "floor":"10–39", "unitQty":30,
     "kitchens":"All unit-07 (30 floors)", "v1":"—", "v2":"—",
     "loadDate":"Jun 30, 2027", "shipDate":"—", "vessel":"—", "etaChs":"—",
     "railDate":"—", "etaNash":"—", "delivery":"Jul 30, 2027", "ordered":False, "confirmed":False,
     "code":"—", "delivTime":"—"},
    {"num":"ECT", "status":"PROJ", "week":"Wk 27", "floor":"—", "unitQty":4,
     "kitchens":"—", "v1":"1004, 2004, 2904, 3604", "v2":"—",
     "loadDate":"—", "shipDate":"—", "vessel":"—", "etaChs":"—",
     "railDate":"—", "etaNash":"—", "delivery":"Aug 6, 2027", "ordered":False, "confirmed":False,
     "code":"—", "delivTime":"—"},
    {"num":"WCT", "status":"PROJ", "week":"Wk 27", "floor":"—", "unitQty":6,
     "kitchens":"1611, 2511, 3311", "v1":"1612, 2512, 3312", "v2":"—",
     "loadDate":"—", "shipDate":"—", "vessel":"—", "etaChs":"—",
     "railDate":"—", "etaNash":"—", "delivery":"Aug 6, 2027", "ordered":False, "confirmed":False,
     "code":"—", "delivTime":"—"},
]

print(f"  → {len(CONTAINERS)} containers total (including BH/ECT/WCT)")

floor_to_ctn    = {c["floor"]: c["num"]    for c in CONTAINERS if isinstance(c["floor"], int)}
floor_to_status = {c["floor"]: c["status"] for c in CONTAINERS if isinstance(c["floor"], int)}

# ─────────────────────────────────────────────────────────────
# 2.  BUILDING MATRIX → EXCEPTIONS
# ─────────────────────────────────────────────────────────────

# The sheet is a floor-by-unit grid: each floor is a 4-row block —
#   row 0: floor number in col B, unit IDs in cols D..O (units 01-12)
#   row 1: "K"  in col C, kitchen status per unit
#   row 2: "V1" in col C, vanity-1 status per unit
#   row 3: "V2" in col C, vanity-2 status per unit ("--" = unit has no V2)
# A unit becomes an "exception" when its component status differs from the
# status the dashboard would otherwise assume (its floor's container status;
# for unit-07 kitchens, the BH container's status).
ws2 = wb[SH_BLD]
EXCEPTIONS = {}
COL_FLOOR, COL_LABEL, COL_U1, COL_U12 = 2, 3, 4, 15   # B, C, D..O

bh_status = next((c["status"] for c in CONTAINERS if c["num"] == "BH"), "PROJ")

def grid_status(v):
    """Grid cell → normalized status code, or None for blank/'--' (no component)."""
    if v is None: return None
    s = str(v).strip()
    if not s or s in ("--", "—"): return None
    return norm_status(s)

blocks_found = 0
for r in range(1, ws2.max_row + 1):
    fv = ws2.cell(r, COL_FLOOR).value
    if not (isinstance(fv, (int, float)) and 10 <= fv <= 39):
        continue
    floor = int(fv)
    # confirm the block shape: next rows labeled K / V1 / V2 in col C
    labels = [str(ws2.cell(r + i, COL_LABEL).value or "").strip() for i in (1, 2, 3)]
    if labels[0] != "K":
        continue
    blocks_found += 1
    main_status = floor_to_status.get(floor)
    for col in range(COL_U1, COL_U12 + 1):
        unit_raw = ws2.cell(r, col).value
        if not isinstance(unit_raw, (int, float)):
            continue
        unit_id = int(unit_raw)
        is_bh   = (unit_id % 100 == 7)
        kS  = grid_status(ws2.cell(r + 1, col).value)
        v1S = grid_status(ws2.cell(r + 2, col).value) if labels[1] == "V1" else None
        v2S = grid_status(ws2.cell(r + 3, col).value) if labels[2] == "V2" else None
        exc = {}
        k_default = bh_status if is_bh else main_status
        if kS  is not None and kS  != k_default:    exc["kStatus"]  = kS
        if v1S is not None and v1S != main_status:  exc["v1Status"] = v1S
        if v2S is not None and v2S != main_status:  exc["v2Status"] = v2S
        if exc:
            EXCEPTIONS[unit_id] = exc

if blocks_found:
    print(f"  → {blocks_found} floor blocks scanned · {len(EXCEPTIONS)} exception units detected")
else:
    print("  ! Building matrix grid not recognized — exceptions left empty")

# ─────────────────────────────────────────────────────────────
# 3.  ROJ DATES
# ─────────────────────────────────────────────────────────────

ROJ = {}
if SH_ROJ is None:
    print("  → ROJ sheet not present — floor-ready dates left empty")
else:
    for row in wb[SH_ROJ].iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 6:
            continue
        floor_raw = row[2]
        if not isinstance(floor_raw, (int, float)):
            continue
        floor = int(floor_raw)
        if floor < 10 or floor > 50:
            continue
        window_raw  = row[5]
        install_raw = row[7] if len(row) > 7 else None
        window_dt   = to_dt(window_raw)
        install_dt  = to_dt(install_raw)
        window      = fmt_date(window_raw)
        install     = "—" if (install_dt and window_dt and install_dt < window_dt) else fmt_date(install_raw)
        if window != "—":
            ROJ[floor] = {"window": window, "install": install}
    print(f"  → ROJ dates parsed for {len(ROJ)} floors")

# ─────────────────────────────────────────────────────────────
# 4.  TRANSIT PERFORMANCE  (sheet "dashboard in progress", C5:D10)
# ─────────────────────────────────────────────────────────────

TRANSIT_PERF = []
SH_DASH = "1010 Dashboard"
if SH_DASH in wb.sheetnames:
    for row in wb[SH_DASH].iter_rows(min_row=12, max_row=16, min_col=2, max_col=4, values_only=True):
        label = clean(row[0]);  value = clean(row[2])
        if label not in ("—", "", "None"):
            TRANSIT_PERF.append({"label": label, "value": value})
    print(f"  → Transit performance: {len(TRANSIT_PERF)} rows from '{SH_DASH}' B12:D16")
else:
    print(f"  ! Sheet '{SH_DASH}' not found — transit performance left empty")

# ─────────────────────────────────────────────────────────────
# 5.  BUILD JS DATA STRINGS
# ─────────────────────────────────────────────────────────────

def js_containers():
    lines = []
    for c in CONTAINERS:
        num = c['num'] if isinstance(c['num'], str) else f"{c['num']:2d}"
        floor = f"'{c['floor']}'" if isinstance(c['floor'], str) else str(c['floor'])
        lines.append(
            f"  {{ num:{js_str(c['num']) if isinstance(c['num'],str) else num}, "
            f"status:{js_str(c['status'])}, week:{js_str(c['week'])}, "
            f"floor:{floor}, unitQty:{c['unitQty']}, "
            f"kitchens:{js_str(c['kitchens'])}, v1:{js_str(c['v1'])}, v2:{js_str(c['v2'])}, "
            f"loadDate:{js_str(c['loadDate'])}, shipDate:{js_str(c['shipDate'])}, "
            f"vessel:{js_str(c['vessel'])}, etaChs:{js_str(c['etaChs'])}, "
            f"railDate:{js_str(c['railDate'])}, etaNash:{js_str(c['etaNash'])}, "
            f"delivery:{js_str(c['delivery'])}, ordered:{'true' if c['ordered'] else 'false'}, "
            f"code:{js_str(c.get('code','—'))}, delivTime:{js_str(c.get('delivTime','—'))} }}"
        )
    return "[\n" + ",\n".join(lines) + "\n]"

def js_roj():
    if not ROJ:
        return "{\n\n}"
    lines = [f"  {floor}: {{ window:{js_str(r['window'])}, install:{js_str(r['install'])} }}"
             for floor in sorted(ROJ.keys())]
    return "{\n" + ",\n".join(lines) + "\n}"

def js_exceptions():
    if not EXCEPTIONS:
        return "{}"
    lines = []
    for uid in sorted(EXCEPTIONS.keys()):
        exc = EXCEPTIONS[uid]
        parts = []
        for key in ("kCtnr","kStatus","v1Ctnr","v1Status","v2Ctnr","v2Status"):
            if key in exc:
                parts.append(f"{key}:{js_str(exc[key])}")
        lines.append(f"  {uid}: {{ {', '.join(parts)} }}")
    return "{\n" + ",\n".join(lines) + "\n}"

def js_transit_perf():
    if not TRANSIT_PERF:
        return "[]"
    items = [f"  {{label:{js_str(r['label'])}, value:{js_str(r['value'])}}}" for r in TRANSIT_PERF]
    return "[\n" + ",\n".join(items) + "\n]"

# ─────────────────────────────────────────────────────────────
# 6.  INJECT DATA INTO EXISTING HTML  (layout untouched)
# ─────────────────────────────────────────────────────────────

generated_on = datetime.now().strftime("%b %#d, %Y at %#I:%M %p")

if UNKNOWN_STATUSES:
    print(f"\n  ! WARNING: unrecognized status code(s) found in Excel: {sorted(UNKNOWN_STATUSES)}")
    print(f"    They are shown on the dashboard exactly as written (uncolored).")
    print(f"    If one is a typo, fix it in the Excel file and rerun. If it's a real")
    print(f"    new status, add it to STATUS_NORM in this script to give it a color.")

print(f"\nUpdating  {os.path.basename(OUTPUT_HTML)} …")
with open(OUTPUT_HTML, "r", encoding="utf-8") as f:
    html = f.read()

# -- Replace data block (everything from "// ── DATA" up to "const V2_UNITS") --
DATA_START = "// ── DATA"
DATA_END   = "const V2_UNITS"

if DATA_START not in html:
    sys.exit(f"\nERROR: Could not find '{DATA_START}' marker in the HTML file.\n"
             f"Make sure the HTML file has not been manually edited to remove that comment.\n")
if DATA_END not in html:
    sys.exit(f"\nERROR: Could not find '{DATA_END}' marker in the HTML file.\n")

start_idx = html.index(DATA_START)
end_idx   = html.index(DATA_END)

new_data_block = (
    f"// ── DATA (auto-generated {generated_on}) ─────────────────────\n"
    f"const CONTAINERS = {js_containers()};\n\n"
    f"const ROJ = {js_roj()};\n\n"
    f"const EXCEPTIONS = {js_exceptions()};\n\n"
    f"const TRANSIT_PERF = {js_transit_perf()};\n\n"
)

html = html[:start_idx] + new_data_block + html[end_idx:]

# -- Update the "Updated ..." timestamp in the subtitle --
html = re.sub(r'<em>Updated [^<]+</em>', f'<em>Updated {generated_on}</em>', html)

with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
    f.write(html)

size_kb = os.path.getsize(OUTPUT_HTML) / 1024
print(f"  → {size_kb:.1f} KB written  ·  {len(CONTAINERS)} containers  ·  {len(EXCEPTIONS)} exceptions")

# -- Write index.html redirect so GitHub Pages root URL works --
INDEX_HTML = os.path.join(SITE_DIR, "index.html")
with open(INDEX_HTML, "w", encoding="utf-8") as f:
    f.write(
        '<!DOCTYPE html>\n'
        '<html lang="en">\n'
        '<head>\n'
        '  <meta charset="utf-8">\n'
        '  <meta http-equiv="refresh" content="0; url=10C_Container_Dashboard.html">\n'
        '  <link rel="canonical" href="10C_Container_Dashboard.html">\n'
        '  <script>window.location.replace("10C_Container_Dashboard.html");</script>\n'
        '</head>\n'
        '<body>\n'
        '  <a href="10C_Container_Dashboard.html">Click here if not redirected automatically.</a>\n'
        '</body>\n'
        '</html>\n'
    )
print(f"  → index.html redirect written")

# ─────────────────────────────────────────────────────────────
# 7.  GIT PUSH
# ─────────────────────────────────────────────────────────────

if AUTO_GIT:
    print("\nPushing to GitHub …")
    def run_git(*args):
        result = subprocess.run(["git"] + list(args), cwd=SITE_DIR,
                                capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip())
        return result.stdout.strip()

    try:
        # Add HTML + any logo files present in the folder
        files_to_add = ["10C_Container_Dashboard.html", "index.html", "config.js"]
        for logo in ["logo-lion.png", "logo-210.png", "logo-dandamudi.png"]:
            if os.path.exists(os.path.join(SITE_DIR, logo)):
                files_to_add.append(logo)
        run_git("add", *files_to_add)
        msg = f"Dashboard update {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        try:
            run_git("commit", "-m", msg)
        except RuntimeError as ce:
            if "nothing to commit" not in str(ce):
                raise
            print("  → No changes to commit — HTML already up to date.")
        run_git("push", "--force", "-u", "origin", "main")
        print(f"  → Pushed: \"{msg}\"")
        print("  → GitHub Pages will refresh in ~30–60 seconds.")
    except RuntimeError as e:
        err = str(e)
        if "git: command not found" in err or ("'git'" in err and "not recognized" in err.lower()):
            print("  ! Git not found. Install Git for Windows from https://git-scm.com")
        else:
            print(f"  ! Git error: {err}")

# ─────────────────────────────────────────────────────────────
# DONE
# ─────────────────────────────────────────────────────────────

print(f"""
{'─'*56}
  Done!
  {len(CONTAINERS)} containers  ·  {len(EXCEPTIONS)} exceptions  ·  {len(ROJ)} ROJ floors
  File:  {OUTPUT_HTML}
{'─'*56}
""")

if sys.platform == "win32":
    input("  Press Enter to close ...")
